import config
import openpyxl
from openpyxl import Workbook
from app.enums import OrdersXlsxSheet


def add_to_xls(sheet: OrdersXlsxSheet, data: list[list[str]]) -> None:
    try:
        workbook = openpyxl.load_workbook(config.XLS_FILE_PATH)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)
        orders_sheet = workbook.create_sheet(OrdersXlsxSheet.ORDERS)
        orders_sheet.append(
            [
                "ID пользователя",
                "Telegram ID",
                "ID заказа",
                "Сумма",
                "Адрес доставки",
                "Дата покупки",
            ]
        )

        details_sheet = workbook.create_sheet(OrdersXlsxSheet.DETAILS)
        details_sheet.append(
            [
                "ID пользователя",
                "Telegram ID",
                "ID заказа",
                "ID товара",
                "Наименование товара",
                "Цена за единицу",
                "Кол-во товара",
                "Сумма",
                "Дата покупки",
            ]
        )

    sheet = workbook[sheet]

    for d in data:
        sheet.append(d)
    workbook.save(config.XLS_FILE_PATH)
